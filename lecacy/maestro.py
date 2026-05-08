
import socket
import threading
import time
import openpyxl 
import os
import sqlite3 
import shutil
import uuid
import platform
import subprocess
import atexit
import json
import secrets
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, redirect, make_response, send_file, render_template, session
from werkzeug.security import generate_password_hash, check_password_hash
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

# Cargar configuración desde .env junto al script.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(BASE_DIR, ".env")

def load_env_file(path):
    if load_dotenv is not None:
        load_dotenv(path)
        return
    if not os.path.exists(path):
        return
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

load_env_file(ENV_FILE)

# Imports opcionales para hardware (Pi4). Si faltan, la app arranca igual.
try:
    import board
    import adafruit_dht
except Exception:
    board = None
    adafruit_dht = None

try:
    from luma.core.interface.serial import i2c as luma_i2c
    from luma.core.render import canvas
    from luma.oled.device import sh1106
except Exception:
    luma_i2c = None
    canvas = None
    sh1106 = None

try:
    from PIL import ImageFont
except Exception:
    ImageFont = None

def env_bool(name, default=False):
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in ("1", "true", "yes", "on")

def env_str(name, default=""):
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip()

PUERTO_WEB = int(os.getenv("PUERTO_WEB", 5000))
PUERTO_UDP = int(os.getenv("PUERTO_UDP", 8081))
MAX_CLIENTS = int(os.getenv("MAX_CLIENTS", 10))
SESSION_TIMEOUT_SECONDS = int(os.getenv("SESSION_TIMEOUT_SECONDS", 30 * 60))

db_name_value = os.getenv("DB_NAME", "casino_data.db")
DB_NAME = db_name_value if os.path.isabs(db_name_value) else os.path.join(BASE_DIR, db_name_value)

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY") or uuid.uuid4().hex
app.permanent_session_lifetime = timedelta(seconds=SESSION_TIMEOUT_SECONDS)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=os.getenv("SESSION_COOKIE_SAMESITE", "Lax"),
    SESSION_COOKIE_SECURE=env_bool("SESSION_COOKIE_SECURE", False),
)

# ESCUDO 2: Candado para evitar colisiones entre la Web y los Sockets
db_lock = threading.RLock()

clientes_tcp = [None] * MAX_CLIENTS
ids_clientes = [""] * MAX_CLIENTS
ips_esclavos =["0.0.0.0"] * MAX_CLIENTS
macs_esclavos = ["Desconocida"] * MAX_CLIENTS
nombres_esclavos =["Maquina " + str(i) for i in range(MAX_CLIENTS)]
ultimos_eventos = ["-"] * MAX_CLIENTS
ultimo_monto = ["0"] * MAX_CLIENTS
ultimo_usuario = ["Desconocido"] * MAX_CLIENTS
ultimo_operation_id = [""] * MAX_CLIENTS
ultimo_pong = [0.0] * MAX_CLIENTS
ultima_conexion = [0.0] * MAX_CLIENTS
ultima_desconexion = [0.0] * MAX_CLIENTS
ultima_sas_desconexion =[0.0] * MAX_CLIENTS
sas_conectado = [False] * MAX_CLIENTS
ultimos_contadores = [""] * MAX_CLIENTS
heartbeat_info =[{"uptime": "-", "firmware": "-", "ultimo_error": "-"} for _ in range(MAX_CLIENTS)]

# Configuración opcional de hardware (Pi4)
ENABLE_OLED = env_bool("ENABLE_OLED", True)
ENABLE_DHT = env_bool("ENABLE_DHT", True)
I2C_PORT = int(os.getenv("I2C_PORT", 1))
I2C_ADDRESS = int(os.getenv("I2C_ADDRESS", "0x3C"), 16)
DHT_PIN_NAME = env_str("DHT_PIN", "D17")
DHT_READ_INTERVAL = int(os.getenv("DHT_READ_INTERVAL", 3))
OLED_REFRESH = float(os.getenv("OLED_REFRESH", 1.0))

# Estado global de clima / OLED
_clima_temp = None
_clima_hum = None
_clima_last_read = 0.0
_dht_device = None
_oled_device = None

usuarios_activos = {}
ultimo_cambio_db = time.time()
ultimo_backup = 0.0

BACKUP_DIR = os.path.join(BASE_DIR, "backups")
ROLES_VALIDOS = ['admin', 'supervisor', 'operador', 'consulta']
MAX_MONTO_ROL = {
    "admin": 1000000.0,
    "supervisor": 250000.0,
    "operador": 50000.0,
    "consulta": 0.0
}
MAX_DIARIO_ROL = {
    "admin": 5000000.0,
    "supervisor": 1000000.0,
    "operador": 250000.0,
    "consulta": 0.0
}
ROLES_VALIDOS = ["admin", "supervisor", "operador", "consulta"]

# ====================================================================
# BASE DE DATOS
# ====================================================================
def init_db():
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("PRAGMA journal_mode=WAL")
        c.execute("PRAGMA busy_timeout=5000")
        c.execute('''CREATE TABLE IF NOT EXISTS transacciones (id INTEGER PRIMARY KEY AUTOINCREMENT, fecha_hora TEXT, usuario TEXT, id_maquina TEXT, ip TEXT, monto REAL)''')
        c.execute('''CREATE TABLE IF NOT EXISTS usuarios (username TEXT PRIMARY KEY, password TEXT, rol TEXT, nombre_real TEXT, password_temporal INTEGER DEFAULT 0)''')
        c.execute("PRAGMA table_info(usuarios)")
        cols_usuarios = [row[1] for row in c.fetchall()]
        if 'password_temporal' not in cols_usuarios:
            c.execute("ALTER TABLE usuarios ADD COLUMN password_temporal INTEGER DEFAULT 0")
        c.execute('''CREATE TABLE IF NOT EXISTS maquinas (id_esclavo TEXT PRIMARY KEY, nombre TEXT)''')
        c.execute('''CREATE TABLE IF NOT EXISTS logs_slot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            slot INTEGER,
            id_esclavo TEXT,
            nombre_maquina TEXT,
            ip TEXT,
            tipo TEXT,
            usuario TEXT,
            detalle TEXT,
            monto REAL,
            contadores TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS acciones_admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            usuario TEXT,
            accion TEXT,
            detalle TEXT,
            ip TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS limites_usuario (
            username TEXT PRIMARY KEY,
            max_operacion REAL,
            max_diario REAL
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS system_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            tipo TEXT,
            detalle TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS solicitudes_limite (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            usuario TEXT,
            monto REAL,
            estado TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS limites_rol (
            rol TEXT PRIMARY KEY,
            max_operacion REAL,
            max_diario REAL
        )''')
        
        c.execute("SELECT count(*) FROM limites_rol")
        if c.fetchone()[0] == 0:
            for rol, m_op in MAX_MONTO_ROL.items():
                m_dia = MAX_DIARIO_ROL.get(rol, 0.0)
                c.execute("INSERT INTO limites_rol VALUES (?, ?, ?)", (rol, m_op, m_dia))

        c.execute("PRAGMA table_info(usuarios)")
        if "nombre_real" not in [col[1] for col in c.fetchall()]:
            c.execute("ALTER TABLE usuarios ADD COLUMN nombre_real TEXT DEFAULT 'Usuario Nuevo'")

        c.execute("PRAGMA table_info(logs_slot)")
        if "operation_id" not in [col[1] for col in c.fetchall()]:
            c.execute("ALTER TABLE logs_slot ADD COLUMN operation_id TEXT")

        c.execute("CREATE INDEX IF NOT EXISTS idx_transacciones_fecha ON transacciones(fecha_hora)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_transacciones_usuario ON transacciones(usuario)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_transacciones_maquina ON transacciones(id_maquina)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_logs_slot_id_esclavo ON logs_slot(id_esclavo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_logs_slot_slot ON logs_slot(slot)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_logs_slot_tipo ON logs_slot(tipo)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_logs_slot_operation ON logs_slot(operation_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_system_logs_tipo ON system_logs(tipo)")

        seed_initial_users(c)
        conn.commit()
        conn.close()

def registrar_historial(id_maq, ip, monto, usr_id):
    global ultimo_cambio_db
    try:
        # ESCUDO 3: Limpiamos la variable para que no crashee si viene basura
        mnt_limpio = str(monto).replace("$", "").replace(",", "").strip()
        mn = float(mnt_limpio) if mnt_limpio else 0.0
        fh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute("SELECT nombre_real FROM usuarios WHERE username=?", (usr_id,))
            res = c.fetchone()
            nom = res[0] if res else str(usr_id)
            c.execute("INSERT INTO transacciones (fecha_hora, usuario, id_maquina, ip, monto) VALUES (?, ?, ?, ?, ?)", (fh, nom, id_maq, ip, mn))
            conn.commit()
            conn.close()
            
        ultimo_cambio_db = time.time() 
    except Exception as e: 
        print(f"[-] Error Critico DB Historial: {e}")

def registrar_log_slot(slot, tipo, detalle, usuario=None, monto=None, contadores=None, operation_id=None):
    global ultimo_cambio_db
    try:
        if slot < 0 or slot >= MAX_CLIENTS:
            return
        fh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        id_esclavo = ids_clientes[slot] if ids_clientes[slot] else ""
        nombre = nombres_esclavos[slot] if nombres_esclavos[slot] else f"Maquina {slot}"
        ip = ips_esclavos[slot] if ips_esclavos[slot] else "0.0.0.0"
        usuario_final = usuario if usuario else "-"
        monto_final = None
        if monto not in (None, ""):
            try:
                monto_final = float(str(monto).replace("$", "").replace(",", "").strip())
            except (TypeError, ValueError):
                monto_final = None

        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute('''INSERT INTO logs_slot
                (fecha_hora, slot, id_esclavo, nombre_maquina, ip, tipo, usuario, detalle, monto, contadores, operation_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (fh, slot, id_esclavo, nombre, ip, tipo, usuario_final, detalle, monto_final, contadores, operation_id))
            conn.commit()
            conn.close()
        ultimo_cambio_db = time.time()
    except Exception as e:
        print(f"[-] Error DB Log Slot: {e}")

def parsear_contadores(raw):
    contadores =[]
    etiquetas = {
        "In": ("Coin In (Apuestas)", True),
        "Out": ("Coin Out (Premios)", True),
        "Drop": ("Drop (Billetero)", True),
        "Canc": ("Cancelled (Pagos manuales)", True),
        "Cred": ("Creditos Disponibles", True),
        "Jack": ("Jackpots (Acumulados)", True),
        "Bill": ("Billetes Aceptados", True),
        "Juego": ("Partidas Jugadas", False)
    }
    for parte in str(raw or "").split():
        if ":" not in parte:
            continue
        clave, valor = parte.split(":", 1)
        nombre, es_moneda = etiquetas.get(clave, (clave, True))
        valor_mostrado = valor
        try:
            numero = float(valor)
            valor_mostrado = round(numero / 100, 2) if es_moneda else int(numero)
        except (TypeError, ValueError):
            pass
        contadores.append((clave, nombre, valor, valor_mostrado))
    return contadores

def nombre_archivo_seguro(valor):
    limpio = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(valor or ""))
    return limpio[:60] if limpio else "sin_id"

def parse_fecha_log(valor):
    try:
        return datetime.strptime(valor, "%d/%m/%Y %H:%M:%S")
    except (TypeError, ValueError):
        return None

def obtener_auditoria_cargas_slot(slot=None, id_esclavo=""):
    where =["tipo IN ('CARGA_ENVIADA', 'CARGA_CONFIRMADA', 'ERROR')"]
    params =[]
    if id_esclavo:
        where.append("id_esclavo=?")
        params.append(id_esclavo)
    elif slot is not None and slot >= 0:
        where.append("slot=?")
        params.append(slot)

    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute('''SELECT id, fecha_hora, slot, id_esclavo, nombre_maquina, ip, tipo, usuario, detalle, monto, operation_id
                     FROM logs_slot WHERE ''' + " AND ".join(where) + ''' ORDER BY id ASC''', tuple(params))
        rows = c.fetchall()
        conn.close()

    enviados =[r for r in rows if r[6] == "CARGA_ENVIADA"]
    confirmados =[r for r in rows if r[6] == "CARGA_CONFIRMADA"]
    errores =[r for r in rows if r[6] == "ERROR"]
    resultado =[]

    for idx, env in enumerate(enviados):
        next_id = enviados[idx + 1][0] if idx + 1 < len(enviados) else None
        env_id, fecha, slot_row, id_row, nombre, ip, _, usuario, detalle, monto, op_id = env
        confirmado = None
        for conf in confirmados:
            if conf[0] <= env_id:
                continue
            if next_id is not None and conf[0] >= next_id:
                continue
            if op_id and conf[10] == op_id:
                confirmado = conf
                break
            if not op_id and str(conf[7] or "") == str(usuario or "") and float(conf[9] or 0) == float(monto or 0):
                confirmado = conf
                break
        error = None
        for err in errores:
            if err[0] <= env_id:
                continue
            if next_id is not None and err[0] >= next_id:
                continue
            error = err
            break

        estado = "Exitoso" if confirmado else "Fallido"
        detalle_estado = confirmado[8] if confirmado else (error[8] if error else "Sin confirmacion registrada")
        resultado.append({
            "fecha": fecha,
            "slot": slot_row,
            "id_esclavo": id_row,
            "maq": nombre,
            "ip": ip,
            "usr": usuario,
            "monto": monto,
            "estado": estado,
            "detalle": detalle_estado,
            "operation_id": op_id or "-"
        })

    return list(reversed(resultado))

def calcular_dashboard_cargas(cargas):
    ahora = datetime.now()
    total_hist = total_mes = total_24 = 0.0
    ultimo_usuario = "-"
    for carga in cargas:
        if carga.get("estado") != "Exitoso":
            continue
        monto = float(carga.get("monto") or 0)
        fecha = parse_fecha_log(carga.get("fecha"))
        total_hist += monto
        if fecha:
            if (ahora - fecha).total_seconds() <= 30 * 24 * 3600:
                total_mes += monto
            if (ahora - fecha).total_seconds() <= 24 * 3600:
                total_24 += monto
        if ultimo_usuario == "-" and carga.get("usr"):
            ultimo_usuario = carga.get("usr")
    return {
        "historico": round(total_hist, 2),
        "mes": round(total_mes, 2),
        "dia": round(total_24, 2),
        "usuario": ultimo_usuario
    }

def inicio_periodo(nombre):
    ahora = datetime.now()
    if nombre == "1h":
        return ahora.timestamp() - 3600
    if nombre == "24h":
        return ahora.timestamp() - (24 * 3600)
    if nombre == "semana":
        return ahora.timestamp() - (7 * 24 * 3600)
    if nombre == "mes":
        return ahora.timestamp() - (30 * 24 * 3600)
    return None

def fecha_en_rango(fecha_texto, rango="", mes=""):
    fecha = parse_fecha_log(fecha_texto)
    if not fecha:
        return False
    if mes:
        return fecha.strftime("%Y-%m") == mes
    inicio = inicio_periodo(rango)
    if inicio is None:
        return True
    return fecha.timestamp() >= inicio

def formato_fecha_ts(ts):
    if not ts:
        return "-"
    try:
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M:%S")
    except (TypeError, ValueError, OSError, OverflowError):
        return "-"

def formato_duracion(segundos):
    try:
        segundos = int(max(0, segundos))
    except (TypeError, ValueError):
        return "-"
    if segundos < 60:
        return f"{segundos}s"
    minutos = segundos // 60
    if minutos < 60:
        return f"{minutos}m"
    horas = minutos // 60
    if horas < 24:
        return f"{horas}h {minutos % 60}m"
    dias = horas // 24
    return f"{dias}d {horas % 24}h"

def normalizar_monto(valor):
    limpio = str(valor or "").replace("$", "").replace(",", "").strip()
    if not limpio:
        return None
    try:
        monto = float(limpio)
    except (TypeError, ValueError):
        return None
    if monto <= 0:
        return None
    return monto

def normalizar_texto_corto(valor, max_len=80):
    return str(valor or "").strip()[:max_len]

def generar_password_temporal(length=4):
    return "".join(secrets.choice("0123456789") for _ in range(length))

def obtener_usuario_login(username):
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT username, password, rol, nombre_real, COALESCE(password_temporal, 0) FROM usuarios WHERE username=?", (username,))
        row = c.fetchone()
        conn.close()
    return row

def iniciar_sesion_usuario(username):
    sid = uuid.uuid4().hex
    datos = datos_sesion_request(request)
    session.permanent = True
    session['usuario'] = username
    session['sid'] = sid
    session['last_activity'] = time.time()
    usuarios_activos[username] = {"sid": sid, "last": time.time(), "ip": datos["ip"], "ua": datos["ua"]}

def log_runtime_warning(contexto, error):
    print(f"[-] {contexto}: {error}")

def parsear_limite_decimal(valor):
    try:
        numero = float(str(valor or "0").replace(',', '.').strip())
    except Exception:
        return None
    if numero < 0:
        return None
    return numero

def validar_fecha_hora_sistema(valor):
    valor = str(valor or "").strip()
    if not valor:
        return None
    try:
        dt = datetime.strptime(valor, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def obtener_nombre_usuario(username):
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT nombre_real FROM usuarios WHERE username=?", (username,))
        row = c.fetchone()
        conn.close()
    return row[0] if row else username

def obtener_limites_usuario(username, rol):
    max_op, max_dia = 0.0, 0.0
    try:
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            # Primero intentamos limites por usuario
            c.execute("SELECT max_operacion, max_diario FROM limites_usuario WHERE username=?", (username,))
            row = c.fetchone()
            if row:
                max_op = float(row[0]) if row[0] is not None else 0.0
                max_dia = float(row[1]) if row[1] is not None else 0.0
            else:
                # Si no hay limite por usuario, buscamos el limite por ROL en la DB
                c.execute("SELECT max_operacion, max_diario FROM limites_rol WHERE rol=?", (rol,))
                row_rol = c.fetchone()
                if row_rol:
                    max_op = float(row_rol[0]) if row_rol[0] is not None else 0.0
                    max_dia = float(row_rol[1]) if row_rol[1] is not None else 0.0
            conn.close()
    except Exception as e:
        log_runtime_warning("Fallback limites usuario", e)
        # Fallback a constantes si falla la DB
        max_op = MAX_MONTO_ROL.get(rol, 0.0)
        max_dia = MAX_DIARIO_ROL.get(rol, 0.0)
    return max_op, max_dia

def validar_password_actual(username, password):
    if not username or not password:
        return False
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT password FROM usuarios WHERE username=?", (username,))
        row = c.fetchone()
        conn.close()
    return bool(row and check_password_hash(row[0], password))

def total_usuario_24h(nombre_real):
    total = 0.0
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, monto FROM transacciones WHERE usuario=?", (nombre_real,))
        rows = c.fetchall()
        conn.close()
    limite = datetime.now() - timedelta(hours=24)
    for fecha_txt, monto in rows:
        fecha = parse_fecha_log(fecha_txt)
        if fecha and fecha >= limite:
            total += float(monto or 0)
    return total

def validar_monto_usuario(monto, username, rol):
    if rol == "consulta":
        return False, "Usuario sin permiso para cargar credito"
    max_op, max_dia = obtener_limites_usuario(username, rol)
    
    # 0 significa SIN LIMITE
    if max_op > 0 and monto > max_op:
        return False, f"Monto supera limite por operacion (${max_op:.2f})"
        
    nombre_real = obtener_nombre_usuario(username)
    usado = total_usuario_24h(nombre_real)
    
    if max_dia > 0 and (usado + monto) > max_dia:
        return False, f"Monto supera limite diario (${max_dia:.2f})"
        
    return True, "OK"

def registrar_accion_admin(usuario, accion, detalle="", ip=""):
    try:
        fh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute("INSERT INTO acciones_admin (fecha_hora, usuario, accion, detalle, ip) VALUES (?, ?, ?, ?, ?)",
                      (fh, usuario or "-", accion, detalle, ip or "-"))
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"[-] Error DB Accion Admin: {e}")

def seed_initial_users(cursor):
    cursor.execute("SELECT count(*) FROM usuarios")
    if cursor.fetchone()[0] != 0:
        return

    admin_user = normalizar_texto_corto(env_str("INIT_ADMIN_USER", "admin"), 40) or "admin"
    admin_name = normalizar_texto_corto(env_str("INIT_ADMIN_NAME", "Administrador Principal"), 80) or "Administrador Principal"
    admin_pass = env_str("INIT_ADMIN_PASS")
    generated_admin_pass = False
    if not admin_pass:
        admin_pass = generar_password_temporal()
        generated_admin_pass = True

    cursor.execute(
        "INSERT INTO usuarios (username, password, rol, nombre_real, password_temporal) VALUES (?, ?, 'admin', ?, ?)",
        (admin_user, generate_password_hash(admin_pass), admin_name, 1 if generated_admin_pass else 0),
    )

    if generated_admin_pass:
        mensaje = f"Admin inicial creado: usuario={admin_user} clave_temporal={admin_pass}"
        print(f"[!] {mensaje}")
        cursor.execute(
            "INSERT INTO system_logs (fecha_hora, tipo, detalle) VALUES (?, ?, ?)",
            (datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "INIT_ADMIN_TEMP_PASSWORD", mensaje),
        )

    operador_user = normalizar_texto_corto(env_str("INIT_OPERATOR_USER"), 40)
    operador_pass = env_str("INIT_OPERATOR_PASS")
    operador_name = normalizar_texto_corto(env_str("INIT_OPERATOR_NAME", "Operador de Sala"), 80) or "Operador de Sala"
    if operador_user and operador_pass:
        cursor.execute(
            "INSERT INTO usuarios (username, password, rol, nombre_real, password_temporal) VALUES (?, ?, 'operador', ?, 0)",
            (operador_user, generate_password_hash(operador_pass), operador_name),
        )

def registrar_system_log(tipo, detalle=""):
    try:
        fh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute("INSERT INTO system_logs (fecha_hora, tipo, detalle) VALUES (?, ?, ?)", (fh, tipo, detalle))
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"[-] Error DB System Log: {e}")

# ====================================================================
# HARDWARE: DHT11 + OLED (opcional, fail-safe)
# ====================================================================
def _get_ip_address():
    try:
        out = subprocess.check_output(["hostname", "-I"], text=True).strip()
        if out:
            return out.split()[0]
    except Exception:
        pass
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(2)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        pass
    return "Sin IP"


def _init_dht():
    global _dht_device
    if not ENABLE_DHT or board is None or adafruit_dht is None:
        return False
    try:
        pin = getattr(board, DHT_PIN_NAME, None)
        if pin is None:
            print(f"[-] DHT: pin '{DHT_PIN_NAME}' no encontrado en board")
            return False
        _dht_device = adafruit_dht.DHT11(pin)
        print("[+] DHT11 inicializado")
        return True
    except Exception as e:
        print(f"[-] DHT11 no disponible: {e}")
        _dht_device = None
        return False


def _read_dht_loop():
    global _clima_temp, _clima_hum, _clima_last_read
    if _dht_device is None:
        return
    while True:
        try:
            if time.time() - _clima_last_read >= DHT_READ_INTERVAL:
                t = _dht_device.temperature
                h = _dht_device.humidity
                if t is not None:
                    _clima_temp = t
                if h is not None:
                    _clima_hum = h
                _clima_last_read = time.time()
        except Exception:
            pass
        time.sleep(1)


def _init_oled():
    global _oled_device
    if not ENABLE_OLED or luma_i2c is None or sh1106 is None or canvas is None:
        return False
    try:
        serial = luma_i2c(port=I2C_PORT, address=I2C_ADDRESS)
        _oled_device = sh1106(serial, width=128, height=64)
        print("[+] OLED SH1106 inicializado")
        return True
    except Exception as e:
        print(f"[-] OLED no disponible: {e}")
        _oled_device = None
        return False


def _oled_display_loop():
    if _oled_device is None:
        return
    font_small = None
    font_bold = None
    font_time = None
    if ImageFont is not None:
        try:
            font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
            font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 10)
            font_time = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        except Exception:
            pass
    if font_small is None:
        font_small = ImageFont.load_default()
    if font_bold is None:
        font_bold = ImageFont.load_default()
    if font_time is None:
        font_time = ImageFont.load_default()

    while True:
        try:
            now = datetime.now()
            hora = now.strftime("%H:%M:%S")
            fecha = now.strftime("%d/%m/%Y")
            ip_addr = _get_ip_address()
            temp_txt = f"[T] {_clima_temp}C" if _clima_temp is not None else "[T] --"
            hum_txt = f"[H] {_clima_hum}%" if _clima_hum is not None else "[H] --"

            with canvas(_oled_device) as draw:
                draw.rectangle(_oled_device.bounding_box, outline="white", fill="black")
                draw.text((4, 4), hora, font=font_time, fill="white")
                draw.text((4, 22), fecha, font=font_small, fill="white")
                draw.text((4, 36), temp_txt, font=font_small, fill="white")
                draw.text((68, 36), hum_txt, font=font_small, fill="white")
                draw.text((4, 52), ip_addr[:20], font=font_bold, fill="white")
        except Exception:
            pass
        time.sleep(OLED_REFRESH)


def get_boot_time_text():
    try:
        with open("/proc/stat", "r") as f:
            for line in f:
                if line.startswith("btime "):
                    return formato_fecha_ts(float(line.split()[1]))
    except Exception as e:
        log_runtime_warning("No se pudo leer boot time", e)
    return "-"

def read_text_file(path):
    try:
        with open(path, "r") as f:
            return f.read().strip()
    except OSError:
        return ""

def obtener_info_rtc():
    rtc_devices = []
    rtc_class_dir = "/sys/class/rtc"
    try:
        if os.path.isdir(rtc_class_dir):
            for entry in sorted(os.listdir(rtc_class_dir)):
                if not entry.startswith("rtc"):
                    continue
                name = read_text_file(os.path.join(rtc_class_dir, entry, "name")) or entry
                rtc_devices.append({
                    "id": entry,
                    "name": name,
                    "dev": f"/dev/{entry}",
                    "present": os.path.exists(f"/dev/{entry}"),
                })
    except Exception:
        rtc_devices = []

    if not rtc_devices and os.path.exists("/dev/rtc0"):
        rtc_devices.append({
            "id": "rtc0",
            "name": "RTC",
            "dev": "/dev/rtc0",
            "present": True,
        })

    return {
        "detected": any(device["present"] for device in rtc_devices),
        "devices": rtc_devices,
    }

def obtener_estado_host():
    total, used, free = shutil.disk_usage(BASE_DIR)
    rtc_info = obtener_info_rtc()
    temp_raw = read_text_file("/sys/class/thermal/thermal_zone0/temp")
    temp = "-"
    if temp_raw:
        try:
            temp = f"{float(temp_raw) / 1000:.1f} C"
        except (TypeError, ValueError):
            temp = temp_raw
    uptime = "-"
    up_raw = read_text_file("/proc/uptime")
    if up_raw:
        try:
            uptime = formato_duracion(float(up_raw.split()[0]))
        except (TypeError, ValueError, IndexError) as e:
            log_runtime_warning("No se pudo interpretar uptime", e)
    load = "-"
    try:
        load = " / ".join([f"{x:.2f}" for x in os.getloadavg()])
    except (AttributeError, OSError) as e:
        log_runtime_warning("No se pudo leer load average", e)
    throttled = "-"
    try:
        out = subprocess.check_output(["vcgencmd", "get_throttled"], timeout=2).decode().strip()
        throttled = out
    except (OSError, subprocess.SubprocessError) as e:
        log_runtime_warning("No se pudo leer throttling", e)
    return {
        "host": platform.node() or "-",
        "sistema": platform.platform(),
        "arquitectura": platform.machine(),
        "python": platform.python_version(),
        "uptime": uptime,
        "boot": get_boot_time_text(),
        "temperatura": temp,
        "load": load,
        "disco_total_gb": round(total / (1024 ** 3), 2),
        "disco_usado_gb": round(used / (1024 ** 3), 2),
        "disco_libre_gb": round(free / (1024 ** 3), 2),
        "throttled": throttled,
        "db_mb": round(os.path.getsize(DB_NAME) / (1024 ** 2), 2) if os.path.exists(DB_NAME) else 0,
        "server_time": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "rtc_detected": rtc_info["detected"],
        "rtc_devices": rtc_info["devices"],
    }

@app.route('/api/system_time', methods=['GET', 'POST'])
def api_system_time():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    
    if request.method == 'POST':
        if rol != 'admin': return "Solo el administrador puede cambiar la hora", 403
        nueva_fecha = validar_fecha_hora_sistema(request.form.get('fecha_hora'))
        if not nueva_fecha:
            registrar_accion_admin(usr, "SET_TIME_RECHAZADO", "Formato de fecha invalido", request.remote_addr)
            return "Fecha invalida. Usa YYYY-MM-DD HH:MM:SS", 400
        
        try:
            # Intentamos ajustar la hora del sistema (requiere privilegios de sudo)
            sistema = platform.system().lower()
            if "linux" in sistema:
                # Usamos timedatectl para sincronizar tambien el RTC si existe
                res = subprocess.run(["sudo", "timedatectl", "set-time", nueva_fecha], capture_output=True, text=True)
                if res.returncode != 0:
                    # Fallback a date si timedatectl falla
                    fallback = subprocess.run(["sudo", "date", "-s", nueva_fecha], capture_output=True, text=True)
                    if fallback.returncode != 0:
                        raise RuntimeError((fallback.stderr or fallback.stdout or res.stderr or "No se pudo ajustar la hora").strip())
            elif "windows" in sistema:
                # Fallback para desarrollo en Windows
                fecha, hora = nueva_fecha.split(" ")
                res_fecha = subprocess.run(["date", fecha], shell=True, capture_output=True, text=True)
                res_hora = subprocess.run(["time", hora], shell=True, capture_output=True, text=True)
                if res_fecha.returncode != 0 or res_hora.returncode != 0:
                    raise RuntimeError((res_fecha.stderr or res_hora.stderr or "No se pudo ajustar la hora").strip())
            else:
                return "Sistema operativo no soportado para cambio de hora", 400
            
            registrar_accion_admin(usr, "SET_TIME", f"Nueva hora: {nueva_fecha}", request.remote_addr)
            registrar_system_log("SET_TIME", f"Hora del sistema ajustada manualmente por {usr}: {nueva_fecha}")
            return "Hora actualizada correctamente"
        except Exception as e:
            registrar_system_log("SET_TIME_ERROR", f"No se pudo ajustar la hora: {e}")
            return f"Error al actualizar la hora: {e}", 500

    return jsonify({
        "server_time": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "timezone": str(time.tzname),
        "rtc_detected": obtener_info_rtc()["detected"],
    })

def registrar_arranque_sistema():
    try:
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute("SELECT tipo FROM system_logs ORDER BY id DESC LIMIT 1")
            row = c.fetchone()
        conn.close()
        if row and row[0] not in ("APP_SHUTDOWN",):
            registrar_system_log("POSIBLE_CORTE_ENERGIA", "El proceso anterior no registro apagado limpio")
    except Exception as e:
        log_runtime_warning("No se pudo revisar arranque previo", e)
    registrar_system_log("APP_START", f"Arranque de AG Maestro. Boot host: {get_boot_time_text()}")

def crear_backup_db():
    global ultimo_backup
    try:
        if not os.path.exists(DB_NAME):
            return
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = os.path.join(BACKUP_DIR, f"casino_data_{ts}.db")
        with db_lock:
            shutil.copy2(DB_NAME, destino)
        backups = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) if f.startswith("casino_data_") and f.endswith(".db")])
        for viejo in backups[:-30]:
            try:
                os.remove(viejo)
            except OSError as e:
                log_runtime_warning(f"No se pudo borrar backup viejo {os.path.basename(viejo)}", e)
        ultimo_backup = time.time()
    except Exception as e:
        print(f"[-] Error Backup DB: {e}")

def backup_loop():
    global ultimo_backup
    while True:
        if time.time() - ultimo_backup > 24 * 3600:
            crear_backup_db()
        time.sleep(3600)

def sesion_activa_valida(usuario, sid=None):
    activo = usuarios_activos.get(usuario)
    if not activo:
        return False
    if isinstance(activo, dict):
        if time.time() - float(activo.get("last", 0)) > SESSION_TIMEOUT_SECONDS:
            usuarios_activos.pop(usuario, None)
            return False
        if sid is not None and activo.get("sid") != sid:
            return False
        return True
    if time.time() - float(activo) > SESSION_TIMEOUT_SECONDS:
        usuarios_activos.pop(usuario, None)
        return False
    return sid is None

def datos_sesion_request(req):
    return {
        "ip": req.remote_addr or "",
        "ua": (req.headers.get("User-Agent") or "")[:160]
    }

def misma_terminal_sesion(activo, req):
    if not isinstance(activo, dict):
        return False
    datos = datos_sesion_request(req)
    return activo.get("ip") == datos["ip"] and activo.get("ua") == datos["ua"]

def invalidar_sesion(usuario=None, sid=None):
    if usuario:
        activo = usuarios_activos.get(usuario)
        if sid is None or sesion_activa_valida(usuario, sid):
            usuarios_activos.pop(usuario, None)
        elif not activo:
            usuarios_activos.pop(usuario, None)
    session.clear()

def check_auth(req):
    user_session = session.get('usuario')
    if not user_session: return "none", None, None
    sid = session.get('sid')
    
    # Verificamos si la sesión es válida en nuestro diccionario global
    if not sesion_activa_valida(user_session, sid):
        # Si no es válida por SID pero el usuario existe y no ha pasado el tiempo de timeout,
        # podríamos estar ante una recuperación de sesión o un error de consistencia.
        # Por seguridad, si el SID no coincide, invalidamos.
        invalidar_sesion(user_session, sid)
        return "none", None, None

    ahora = time.time()
    last = session.get('last_activity', ahora)
    
    # Verificamos timeout de inactividad
    if ahora - float(last) > SESSION_TIMEOUT_SECONDS:
        invalidar_sesion(user_session, sid)
        return "none", None, None
        
    session['last_activity'] = ahora
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT rol, nombre_real FROM usuarios WHERE username=?", (user_session,))
        row = c.fetchone()
        conn.close()
    
    if row:
        # Actualizamos el timestamp en el diccionario global
        if user_session in usuarios_activos:
            if isinstance(usuarios_activos[user_session], dict):
                usuarios_activos[user_session]["last"] = ahora
                datos = datos_sesion_request(req)
                usuarios_activos[user_session]["ip"] = datos["ip"]
                usuarios_activos[user_session]["ua"] = datos["ua"]
            else:
                # Migración de formato simple a dict si fuera necesario
                datos = datos_sesion_request(req)
                usuarios_activos[user_session] = {"sid": sid, "last": ahora, "ip": datos["ip"], "ua": datos["ua"]}
        return row[0], user_session, row[1]

    invalidar_sesion(user_session, sid)
    return "none", None, None

# ====================================================================
# FRONTEND (HTML + JS)
# ====================================================================
# Frontend desacoplado (Fase 1)


# ====================================================================
# RUTAS API
# ====================================================================
@app.route('/')
def index():
    if check_auth(request)[0] == "none": return redirect('/login')
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u, p = request.form.get('u'), request.form.get('p')
        force = request.form.get('force') == '1'
        cambio_temporal = request.form.get('change_temp') == '1'

        row = obtener_usuario_login(u)

        if cambio_temporal:
            nueva_clave = request.form.get('new_password') or ''
            confirmar_clave = request.form.get('confirm_password') or ''
            if not (row and check_password_hash(row[1], p) and int(row[4] or 0) == 1):
                registrar_accion_admin(u, "CAMBIO_TEMPORAL_INVALIDO", "Intento invalido de cambio de clave temporal", request.remote_addr)
                return render_template('login.html', error_msg="La clave temporal ya no es válida. Inicia sesión nuevamente.", prev_u=u)
            if not nueva_clave.strip():
                return render_template('login.html', error_msg="Ingresa una nueva clave.", show_temp_change=True, prev_u=u, prev_p=p)
            if nueva_clave != confirmar_clave:
                return render_template('login.html', error_msg="Las claves no coinciden.", show_temp_change=True, prev_u=u, prev_p=p)
            if nueva_clave == p:
                return render_template('login.html', error_msg="La nueva clave debe ser distinta de la temporal.", show_temp_change=True, prev_u=u, prev_p=p)
            nuevo_hash = generate_password_hash(nueva_clave)
            with db_lock:
                conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
                c.execute("UPDATE usuarios SET password=?, password_temporal=0 WHERE username=?", (nuevo_hash, u))
                conn.commit(); conn.close()
            iniciar_sesion_usuario(u)
            registrar_accion_admin(u, "CAMBIO_CLAVE_TEMPORAL", "Clave temporal renovada al iniciar sesion", request.remote_addr)
            registrar_accion_admin(u, "LOGIN_OK", "Ingreso exitoso luego de cambio de clave temporal", request.remote_addr)
            return redirect('/')

        if row and check_password_hash(row[1], p): 
            if int(row[4] or 0) == 1:
                return render_template('login.html', info_msg="Debes cambiar la clave temporal antes de continuar.", show_temp_change=True, prev_u=u, prev_p=p)
            if sesion_activa_valida(u):
                if not force:
                    registrar_accion_admin(u, "LOGIN_BLOQUEADO", "Sesion simultanea detectada", request.remote_addr)
                    return render_template('login.html', error_msg="Esta cuenta ya tiene una sesión activa.", show_force_btn=True, prev_u=u, prev_p=p)
                else:
                    usuarios_activos.pop(u, None)
                    registrar_accion_admin(u, "LOGIN_FORZADO", "Cierre de sesión previa e ingreso", request.remote_addr)

            iniciar_sesion_usuario(u)
            registrar_accion_admin(u, "LOGIN_OK", "Ingreso exitoso", request.remote_addr)
            return redirect('/')
        
        registrar_accion_admin(u, "LOGIN_FALLIDO", "Usuario o clave invalida", request.remote_addr)
        return render_template('login.html', error_msg="Usuario o clave inválida", prev_u=u)
        
    return render_template('login.html')

@app.route('/logout')
def logout(): 
    cookie = session.get('usuario')
    sid = session.get('sid')
    registrar_accion_admin(cookie, "LOGOUT", "Salida de sesion", request.remote_addr)
    invalidar_sesion(cookie, sid)
    return redirect('/login')

def registrar_apagado_limpio():
    registrar_system_log("APP_SHUTDOWN", "Cierre limpio del proceso AG Maestro")

@app.route('/api/estado')
def estado():
    rol, usr, nom = check_auth(request)
    if rol == "none": return "401", 401
    
    max_op, max_dia = obtener_limites_usuario(usr, rol)
    usado = total_usuario_24h(nom)
    disponible = max(0, max_dia - usado) if max_dia > 0 else "Sin límite"

    esc = [{"slot":i, "nombre":nombres_esclavos[i], "id":ids_clientes[i], "ip":ips_esclavos[i], "mac":macs_esclavos[i], "online":clientes_tcp[i] is not None, "sas":sas_conectado[i], "evento":ultimos_eventos[i], "contadores":ultimos_contadores[i]} for i in range(MAX_CLIENTS) if ids_clientes[i]]
    return jsonify({
        "rol":rol, "user":usr, "nombre_real":nom, 
        "limite_disponible": disponible,
        "esclavos":esc, "db_timestamp":ultimo_cambio_db
    })

@app.route('/api/usuarios/add_extra_credit', methods=['POST'])
def add_extra_credit():
    if check_auth(request)[0] != 'admin': return "403", 403
    usuario_obj = request.form.get('u')
    try:
        monto_extra = float(str(request.form.get('monto', '0')).replace(',', '.'))
        if monto_extra <= 0: return "Monto invalido", 400
    except (TypeError, ValueError): return "Monto invalido", 400
    
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("SELECT rol FROM usuarios WHERE username=?", (usuario_obj,))
        res = c.fetchone()
        if not res: return "Usuario no encontrado", 404
        
        # Obtenemos limite actual
        max_op, max_dia = obtener_limites_usuario(usuario_obj, res[0])
        nuevo_dia = max_dia + monto_extra
        
        c.execute("""INSERT INTO limites_usuario (username, max_operacion, max_diario)
                     VALUES (?,?,?)
                     ON CONFLICT(username) DO UPDATE SET max_diario=excluded.max_diario""",
                  (usuario_obj, max_op, nuevo_dia))
        conn.commit(); conn.close()
    
    registrar_accion_admin(session.get('usuario'), "CREDITO_EXTRA", f"Usuario {usuario_obj} +${monto_extra} (Total: {nuevo_dia})", request.remote_addr)
    return "OK"

@app.route('/api/host_status')
def host_status():
    rol, _, _ = check_auth(request)
    if rol == "none": return "401", 401
    status = obtener_estado_host()
    status["rol"] = rol
    return jsonify(status)

@app.route('/api/clima')
def api_clima():
    rol, _, _ = check_auth(request)
    if rol == "none": return "401", 401
    return jsonify({
        "temperatura": _clima_temp,
        "humedad": _clima_hum,
        "ip": _get_ip_address(),
    })

@app.route('/api/reboot_host', methods=['POST'])
def api_reboot_host():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    try:
        sistema = platform.system().lower()
        if "windows" in sistema:
            cmd = ["shutdown", "/r", "/t", "5", "/f"]
        elif "linux" in sistema or "darwin" in sistema:
            cmd = ["shutdown", "-r", "+1"]
        else:
            return "Sistema operativo no soportado para reboot", 400
        subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        registrar_accion_admin(usr, "REBOOT_HOST", f"Host {platform.node()} reiniciado desde web", request.remote_addr)
        registrar_system_log("REBOOT_HOST", f"Reinicio del host solicitado por {usr}")
        return "Reinicio del host programado"
    except Exception as e:
        registrar_system_log("REBOOT_HOST_ERROR", f"No se pudo programar reinicio: {e}")
        return f"No se pudo programar el reinicio: {e}", 500

@app.route('/api/system_logs')
def api_system_logs():
    if check_auth(request)[0] == 'none': return "401", 401
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, tipo, detalle FROM system_logs ORDER BY id DESC LIMIT 200")
        rows = c.fetchall()
        conn.close()
    return jsonify([{"fecha": r[0], "tipo": r[1], "detalle": r[2]} for r in rows])

@app.route('/api/db_info', methods=['POST'])
def api_db_info():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    if not validar_password_actual(usr, request.form.get('password')):
        registrar_accion_admin(usr, "DB_INFO_RECHAZADO", "Password incorrecto", request.remote_addr)
        return "Contraseña incorrecta", 403
    tablas =[]
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name")
        nombres = [r[0] for r in c.fetchall()]
        for nombre in nombres:
            try:
                tabla_segura = '"' + nombre.replace('"', '""') + '"'
                c.execute(f"SELECT COUNT(*) FROM {tabla_segura}")
                tablas.append({"nombre": nombre, "registros": c.fetchone()[0]})
            except Exception:
                tablas.append({"nombre": nombre, "registros": "Error"})
        conn.close()
    backups =[]
    os.makedirs(BACKUP_DIR, exist_ok=True)
    for archivo in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if not (archivo.startswith("casino_data_") and archivo.endswith(".db")):
            continue
        path = os.path.join(BACKUP_DIR, archivo)
        backups.append({
            "archivo": archivo,
            "fecha": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%d/%m/%Y %H:%M:%S"),
            "mb": round(os.path.getsize(path) / (1024 * 1024), 2)
        })
    registrar_accion_admin(usr, "DB_INFO", "Visualizo estado de base de datos", request.remote_addr)
    return jsonify({
        "archivo": os.path.basename(DB_NAME),
        "mb": round(os.path.getsize(DB_NAME) / (1024 * 1024), 2) if os.path.exists(DB_NAME) else 0,
        "ultimo_backup": formato_fecha_ts(ultimo_backup) if ultimo_backup else "-",
        "tablas": tablas,
        "backups": backups[:30]
    })

@app.route('/api/db_restore', methods=['POST'])
def api_db_restore():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    if not validar_password_actual(usr, request.form.get('password')):
        registrar_accion_admin(usr, "DB_RESTORE_RECHAZADO", "Password incorrecto", request.remote_addr)
        return "Contraseña incorrecta", 403
    archivo = os.path.basename(request.form.get('archivo') or "")
    if not (archivo.startswith("casino_data_") and archivo.endswith(".db")):
        return "Backup invalido", 400
    origen = os.path.abspath(os.path.join(BACKUP_DIR, archivo))
    backup_dir_abs = os.path.abspath(BACKUP_DIR)
    if not origen.startswith(backup_dir_abs) or not os.path.exists(origen):
        return "Backup no encontrado", 404
    try:
        crear_backup_db()
        with db_lock:
            shutil.copy2(origen, DB_NAME)
        registrar_accion_admin(usr, "DB_RESTORE", f"Restauro backup {archivo}", request.remote_addr)
        registrar_system_log("DB_RESTORE", f"Base restaurada desde {archivo} por {usr}")
        return "Base restaurada. Reinicia la aplicación si notas datos en cache."
    except Exception as e:
        registrar_system_log("DB_RESTORE_ERROR", f"Error restaurando {archivo}: {e}")
        return f"No se pudo restaurar la base: {e}", 500

@app.route('/api/dashboard_principal')
def dashboard_principal():
    if check_auth(request)[0] == 'none': return "401", 401
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, usuario, monto FROM transacciones")
        rows = c.fetchall()
        conn.close()

    ahora = datetime.now()
    montos = {"hora": 0.0, "dia": 0.0, "mes": 0.0}
    globales = {
        "1h": {"cantidad": 0, "monto": 0.0},
        "24hs": {"cantidad": 0, "monto": 0.0},
        "Ultimo mes": {"cantidad": 0, "monto": 0.0}
    }
    usuarios = {}
    for fecha_txt, usuario, monto in rows:
        fecha = parse_fecha_log(fecha_txt)
        if not fecha:
            continue
        diff = (ahora - fecha).total_seconds()
        valor = float(monto or 0)
        usr_key = usuario or "-"
        if usr_key not in usuarios:
            usuarios[usr_key] = {"cantidad": 0, "monto": 0.0}
        usuarios[usr_key]["cantidad"] += 1
        usuarios[usr_key]["monto"] += valor
        if diff <= 3600:
            montos["hora"] += valor
            globales["1h"]["cantidad"] += 1
            globales["1h"]["monto"] += valor
        if diff <= 24 * 3600:
            montos["dia"] += valor
            globales["24hs"]["cantidad"] += 1
            globales["24hs"]["monto"] += valor
        if diff <= 30 * 24 * 3600:
            montos["mes"] += valor
            globales["Ultimo mes"]["cantidad"] += 1
            globales["Ultimo mes"]["monto"] += valor

    conectados = sum(1 for c in clientes_tcp if c is not None)
    sas_desconectado = sum(1 for i in range(MAX_CLIENTS) if clientes_tcp[i] is not None and not sas_conectado[i])
    por_usuario = [{"usuario": k, "cantidad": v["cantidad"], "monto": round(v["monto"], 2)} for k, v in sorted(usuarios.items(), key=lambda x: x[1]["cantidad"], reverse=True)[:8]]
    detalle_conectados =[]
    detalle_sas_off = []
    for i in range(MAX_CLIENTS):
        if not ids_clientes[i]:
            continue
        online = clientes_tcp[i] is not None
        tiempo_desc = "-"
        if not online:
            tiempo_desc = formato_duracion(time.time() - (ultima_desconexion[i] or ultimo_pong[i] or time.time()))
        elif not sas_conectado[i]:
            tiempo_desc = formato_duracion(time.time() - (ultima_sas_desconexion[i] or ultima_conexion[i] or time.time()))
        detalle = {
            "slot": i,
            "nombre": nombres_esclavos[i],
            "id": ids_clientes[i],
            "ip": ips_esclavos[i],
            "sas": "OK" if sas_conectado[i] else "OFF",
            "ultima_conexion": formato_fecha_ts(ultima_conexion[i] or ultimo_pong[i]),
            "tiempo_desconexion": tiempo_desc,
            "evento": ultimos_eventos[i],
            "uptime": heartbeat_info[i].get("uptime", "-"),
            "firmware": heartbeat_info[i].get("firmware", "-"),
            "ultimo_error": heartbeat_info[i].get("ultimo_error", "-")
        }
        if online:
            detalle_conectados.append(detalle)
            if not sas_conectado[i]:
                detalle_sas_off.append(detalle)
    return jsonify({
        "montos": {k: round(v, 2) for k, v in montos.items()},
        "globales": [{"periodo": k, "cantidad": v["cantidad"], "monto": round(v["monto"], 2)} for k, v in globales.items()],
        "por_usuario": por_usuario,
        "slots": {"conectados": conectados, "sas_desconectado": sas_desconectado},
        "slot_detalle": {"conectados": detalle_conectados, "sas_desconectado": detalle_sas_off}
    })

@app.route('/api/alertas')
def api_alertas():
    if check_auth(request)[0] == 'none': return "401", 401
    ahora = time.time()
    slots_registrados = [i for i in range(MAX_CLIENTS) if ids_clientes[i]]
    slots_offline = sum(1 for i in slots_registrados if clientes_tcp[i] is None)
    sas_off = sum(1 for i in slots_registrados if clientes_tcp[i] is not None and not sas_conectado[i])
    sin_heartbeat = sum(1 for i in slots_registrados if clientes_tcp[i] is not None and ahora - ultimo_pong[i] > 10)

    cargas_fallidas_24h = 0
    limite = datetime.now() - timedelta(hours=24)
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, tipo FROM logs_slot WHERE tipo='ERROR'")
        rows = c.fetchall()
        
        c.execute("SELECT count(*) FROM solicitudes_limite WHERE estado='PENDIENTE'")
        sol_pendientes = c.fetchone()[0]
        conn.close()
    for fecha_txt, tipo in rows:
        fecha = parse_fecha_log(fecha_txt)
        if not fecha or fecha < limite:
            continue
        if tipo == "ERROR":
            cargas_fallidas_24h += 1
    return jsonify({
        "slots_offline": slots_offline,
        "sas_off": sas_off,
        "cargas_fallidas_24h": cargas_fallidas_24h,
        "sin_heartbeat": sin_heartbeat,
        "solicitudes_pendientes": sol_pendientes
    })

@app.route('/api/usuarios_dashboard')
def usuarios_dashboard():
    if check_auth(request)[0] != 'admin': return jsonify([])
    rango = (request.args.get('rango') or "mes").strip()
    inicio = inicio_periodo(rango)
    datos = {}
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, usuario, monto FROM transacciones")
        rows = c.fetchall()
        conn.close()
    for fecha_txt, usuario, monto in rows:
        fecha = parse_fecha_log(fecha_txt)
        if not fecha:
            continue
        if inicio is not None and fecha.timestamp() < inicio:
            continue
        key = usuario or "-"
        if key not in datos:
            datos[key] = {"cantidad": 0, "monto": 0.0}
        datos[key]["cantidad"] += 1
        datos[key]["monto"] += float(monto or 0)
    return jsonify([{"usuario": k, "cantidad": v["cantidad"], "monto": round(v["monto"], 2)} for k, v in sorted(datos.items(), key=lambda x: x[1]["monto"], reverse=True)])

@app.route('/api/perfil/nombre', methods=['POST'])
def chg_name():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor(); c.execute("UPDATE usuarios SET nombre_real=? WHERE username=?", (request.form.get('n'), usr)); conn.commit(); conn.close()
    return "OK"

@app.route('/api/perfil/pass', methods=['POST'])
def chg_pass():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("SELECT password FROM usuarios WHERE username=?", (usr,))
        row = c.fetchone()
        
        if row and check_password_hash(row[0], request.form.get('old')):
            nuevo_hash = generate_password_hash(request.form.get('new'))
            c.execute("UPDATE usuarios SET password=?, password_temporal=0 WHERE username=?", (nuevo_hash, usr))
            conn.commit(); conn.close()
            return "OK"
        conn.close()
    return "Error", 400

@app.route('/api/usuarios/edit', methods=['POST'])
def edit_usr():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    rol_nuevo = normalizar_texto_corto(request.form.get('r'), 20)
    if rol_nuevo not in ROLES_VALIDOS: return "Rol invalido", 400
    usuario_obj = normalizar_texto_corto(request.form.get('u'), 40)
    nombre_real = normalizar_texto_corto(request.form.get('n'), 80)
    max_operacion = parsear_limite_decimal(request.form.get('max_operacion', ''))
    max_diario = parsear_limite_decimal(request.form.get('max_diario', ''))
    if not usuario_obj or not nombre_real:
        return "Usuario y nombre son obligatorios", 400
    if max_operacion is None or max_diario is None:
        return "Limites invalidos", 400
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("UPDATE usuarios SET nombre_real=?, rol=? WHERE username=?", (nombre_real, rol_nuevo, usuario_obj))
        if c.rowcount <= 0:
            conn.close()
            return "Usuario no encontrado", 404
        c.execute("""INSERT INTO limites_usuario (username, max_operacion, max_diario)
                     VALUES (?,?,?)
                     ON CONFLICT(username) DO UPDATE SET max_operacion=excluded.max_operacion, max_diario=excluded.max_diario""",
                  (usuario_obj, max_operacion, max_diario))
        conn.commit(); conn.close()
    registrar_accion_admin(usr, "EDITAR_USUARIO", f"Usuario {usuario_obj} rol {rol_nuevo} limites op:{max_operacion} dia:{max_diario}", request.remote_addr)
    return "OK"

@app.route('/api/usuarios/reset', methods=['POST'])
def rst_pass():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    usuario_obj = normalizar_texto_corto(request.form.get('u'), 40)
    if not usuario_obj:
        return "Usuario invalido", 400
    nueva_clave = generar_password_temporal()
    hash_reset = generate_password_hash(nueva_clave)
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("UPDATE usuarios SET password=?, password_temporal=1 WHERE username=?", (hash_reset, usuario_obj))
        if c.rowcount <= 0:
            conn.close()
            return "Usuario no encontrado", 404
        conn.commit(); conn.close()
    registrar_accion_admin(usr, "RESET_CLAVE", f"Usuario {usuario_obj}", request.remote_addr)
    return f"Clave temporal para {usuario_obj}: {nueva_clave}"

@app.route('/api/usuarios/edit_limits', methods=['POST'])
def edit_usr_limits():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    usuario_obj = normalizar_texto_corto(request.form.get('u'), 40)
    max_operacion = parsear_limite_decimal(request.form.get('max_operacion', '0'))
    max_diario = parsear_limite_decimal(request.form.get('max_diario', '0'))
    if not usuario_obj:
        return "Usuario invalido", 400
    if max_operacion is None or max_diario is None:
        return "Limites invalidos", 400
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("""INSERT INTO limites_usuario (username, max_operacion, max_diario)
                     VALUES (?,?,?)
                     ON CONFLICT(username) DO UPDATE SET max_operacion=excluded.max_operacion, max_diario=excluded.max_diario""",
                  (usuario_obj, max_operacion, max_diario))
        conn.commit(); conn.close()
    registrar_accion_admin(usr, "EDITAR_LIMITES_USUARIO", f"Usuario {usuario_obj} op:{max_operacion} dia:{max_diario}", request.remote_addr)
    return "OK"

@app.route('/api/usuarios')
def ls_usr():
    if check_auth(request)[0] != 'admin': return jsonify([])
    
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT username, rol, nombre_real FROM usuarios")
        rows = c.fetchall()
        conn.close()
        
    users_data =[]
    
    for r in rows:
        u, rol, nom = r[0], r[1], r[2]
        is_online = sesion_activa_valida(u)
        
        # Ahora es seguro llamar a esto porque el candado principal esta liberado (y ademas usamos RLock por seguridad)
        max_op, max_dia = obtener_limites_usuario(u, rol)
        
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10)
            c = conn.cursor()
            c.execute("SELECT monto, fecha_hora, id_maquina FROM transacciones WHERE usuario=? ORDER BY id DESC LIMIT 1", (nom,))
            ult = c.fetchone()
            conn.close()
            
        ult_carga, ult_min, ult_maq = None, 0, ""
        if ult:
            ult_carga = ult[0]
            ult_maq = ult[2]
            try:
                diff = datetime.now() - datetime.strptime(ult[1], "%d/%m/%Y %H:%M:%S")
                ult_min = int(diff.total_seconds() / 60)
            except (TypeError, ValueError):
                ult_min = 0
            
        users_data.append({
            "user": u, "rol": rol, "nombre": nom, "online": is_online,
            "max_operacion": max_op, "max_diario": max_dia,
            "usado_24h": total_usuario_24h(nom),
            "ultima_carga": ult_carga, "ultima_carga_min": ult_min, "ultima_maquina": ult_maq
        })
        
    return jsonify(users_data)

@app.route('/api/usuarios/add', methods=['POST'])
def add_usr():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    rol_nuevo = normalizar_texto_corto(request.form.get('r'), 20)
    if rol_nuevo not in ROLES_VALIDOS: return "Rol invalido", 400
    usuario_obj = normalizar_texto_corto(request.form.get('u'), 40)
    nombre_real = normalizar_texto_corto(request.form.get('n'), 80)
    password_nuevo = request.form.get('p') or ""
    max_operacion = parsear_limite_decimal(request.form.get('max_operacion', '0'))
    max_diario = parsear_limite_decimal(request.form.get('max_diario', '0'))
    if not usuario_obj or not nombre_real:
        return "Usuario y nombre son obligatorios", 400
    if not password_nuevo.strip():
        return "Clave invalida", 400
    if max_operacion is None or max_diario is None:
        return "Limites invalidos", 400

    try:
        hash_pass = generate_password_hash(password_nuevo)
        with db_lock:
            conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO usuarios (username, password, rol, nombre_real, password_temporal) VALUES (?,?,?,?,0)", 
                      (usuario_obj, hash_pass, rol_nuevo, nombre_real))
            c.execute("""INSERT INTO limites_usuario (username, max_operacion, max_diario)
                         VALUES (?,?,?)
                         ON CONFLICT(username) DO UPDATE SET max_operacion=excluded.max_operacion, max_diario=excluded.max_diario""",
                      (usuario_obj, max_operacion, max_diario))
            conn.commit(); conn.close()
        registrar_accion_admin(usr, "CREAR_USUARIO", f"Usuario {usuario_obj} rol {rol_nuevo} limites op:{max_operacion} dia:{max_diario}", request.remote_addr)
    except Exception as e:
        registrar_system_log("ADD_USER_ERROR", f"No se pudo guardar usuario {usuario_obj}: {e}")
        return f"No se pudo guardar el usuario: {e}", 500
    return "OK"

@app.route('/api/usuarios/delete', methods=['POST'])
def del_usr():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    usuario_obj = normalizar_texto_corto(request.form.get('u'), 40)
    if rol != 'admin' or usuario_obj == 'admin': return "403", 403
    if not usuario_obj:
        return "Usuario invalido", 400
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10); c = conn.cursor()
        c.execute("DELETE FROM limites_usuario WHERE username=?", (usuario_obj,))
        c.execute("DELETE FROM usuarios WHERE username=?", (usuario_obj,))
        conn.commit(); conn.close()
    usuarios_activos.pop(usuario_obj, None)
    registrar_accion_admin(usr, "ELIMINAR_USUARIO", f"Usuario {usuario_obj}", request.remote_addr)
    return "OK"

@app.route('/api/comando', methods=['POST'])
def cmd():
    rol, usr, _ = check_auth(request)
    o = normalizar_texto_corto(request.form.get('o'), 64)
    c = normalizar_texto_corto(request.form.get('c'), 64)
    targets_raw = request.form.get('targets', '')
    targets = set(x.strip() for x in targets_raw.split(',') if x.strip())
    if rol == "none": return "401", 401
    if not o or not c:
        return "Comando invalido", 400
    if c == "REBOOT" and rol != "admin": return "403", 403
    if c == "METERS" and rol == "consulta": return "403", 403
    if o == "MULTI" and not targets:
        return "Sin slots seleccionados", 400
    monto_normalizado = None
    if c not in ["REBOOT", "METERS"]:
        monto_normalizado = normalizar_monto(c)
        if monto_normalizado is None:
            registrar_accion_admin(usr, "COMANDO_RECHAZADO", f"Comando/monto invalido: {c}", request.remote_addr)
            return "Monto invalido", 400
        ok, msg = validar_monto_usuario(monto_normalizado, usr, rol)
        if not ok:
            registrar_accion_admin(usr, "CARGA_RECHAZADA", msg, request.remote_addr)
            return msg, 403
        if o == "ALL" or targets:
            destinos = sum(1 for i in range(MAX_CLIENTS) if clientes_tcp[i] and (o == "ALL" or ids_clientes[i] in targets))
            if destinos <= 0:
                return "Sin slots online", 400
            _, max_dia = obtener_limites_usuario(usr, rol)
            usado = total_usuario_24h(obtener_nombre_usuario(usr))
            if usado + (monto_normalizado * destinos) > max_dia:
                msg = f"Carga multiple supera limite diario (${max_dia:.2f})"
                registrar_accion_admin(usr, "CARGA_RECHAZADA", msg, request.remote_addr)
                return msg, 403
        c = str(int(monto_normalizado)) if monto_normalizado.is_integer() else str(monto_normalizado)
    enviados = 0
    for i in range(MAX_CLIENTS):
        if clientes_tcp[i] and (o == "ALL" or ids_clientes[i] == o or ids_clientes[i] in targets):
            try: 
                clientes_tcp[i].sendall((c + "\r\n").encode())
                enviados += 1
                if c not in["REBOOT", "METERS"]:
                    ultimo_monto[i] = c
                    ultimo_usuario[i] = usr
                    operation_id = uuid.uuid4().hex[:12].upper()
                    ultimo_operation_id[i] = operation_id
                    registrar_log_slot(i, "CARGA_ENVIADA", f"OP:{operation_id} Enviando carga por sistema: ${c}", usr, c, operation_id=operation_id)
                
                if c == "REBOOT":
                    ultimos_eventos[i] = "Reiniciando..."
                    registrar_log_slot(i, "EVENTO", "Reinicio solicitado desde web", usr)
                    registrar_accion_admin(usr, "REBOOT_SLOT", f"Slot {i} / {ids_clientes[i]}", request.remote_addr)
                elif c == "METERS": 
                    ultimos_eventos[i] = "Consultando contadores..."
                    ultimos_contadores[i] = "" 
                    registrar_log_slot(i, "EVENTO", "Consulta de contadores solicitada", usr)
                else: ultimos_eventos[i] = f"Enviando ${c}..."
            except Exception as e:
                registrar_log_slot(i, "ERROR", f"Error enviando comando: {e}", usr)
    if enviados <= 0:
        return "Sin slots online", 400
    return f"OK ({enviados} slot/s)"

@app.route('/api/renombrar', methods=['POST'])
def ren():
    if check_auth(request)[0] != 'admin': return "403", 403
    s, n = int(request.form.get('s')), request.form.get('n')
    nombres_esclavos[s] = n
    if ids_clientes[s]: 
        with db_lock:
            conn=sqlite3.connect(DB_NAME, timeout=10); c=conn.cursor(); c.execute("INSERT OR REPLACE INTO maquinas VALUES (?,?)", (ids_clientes[s], n)); conn.commit(); conn.close()
    registrar_accion_admin(session.get('usuario'), "RENOMBRAR_SLOT", f"Slot {s} -> {n}", request.remote_addr)
    return "OK"

@app.route('/api/remove', methods=['POST'])
def rem():
    if check_auth(request)[0] != 'admin': return "403", 403
    s = int(request.form.get('s'))
    if not clientes_tcp[s]: ids_clientes[s], nombres_esclavos[s] = "", f"Maquina {s}"
    registrar_accion_admin(session.get('usuario'), "QUITAR_SLOT", f"Slot {s}", request.remote_addr)
    return "OK"

@app.route('/api/auditoria')
def aud():
    if check_auth(request)[0] == 'none': return "401", 401
    with db_lock:
        conn=sqlite3.connect(DB_NAME, timeout=10); c=conn.cursor(); c.execute('SELECT id, fecha_hora, usuario, id_maquina, ip, monto FROM transacciones ORDER BY id DESC LIMIT 500'); r=c.fetchall(); conn.close()
    return jsonify([{"id":x[0], "fecha":x[1], "usr":x[2], "maq":x[3], "ip":x[4], "monto":x[5]} for x in r])

@app.route('/api/slot_log')
def api_slot_log():
    if check_auth(request)[0] == 'none': return "401", 401
    try:
        slot = int(request.args.get('slot', '-1'))
    except (TypeError, ValueError):
        return jsonify([])
    id_esclavo = (request.args.get('id_esclavo') or "").strip()

    where = []
    params =[]
    if id_esclavo:
        where.append("id_esclavo=?")
        params.append(id_esclavo)
    elif slot >= 0:
        where.append("slot=?")
        params.append(slot)

    sql = '''SELECT fecha_hora, slot, id_esclavo, tipo, usuario, detalle, monto
             FROM logs_slot'''
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC LIMIT 300"

    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute(sql, tuple(params))
        rows = c.fetchall()
        conn.close()

    return jsonify([{
        "fecha": x[0],
        "slot": x[1],
        "id_esclavo": x[2],
        "tipo": x[3],
        "usuario": x[4],
        "detalle": x[5],
        "monto": x[6]
    } for x in rows])

@app.route('/api/slot_cargas')
def api_slot_cargas():
    if check_auth(request)[0] == 'none': return "401", 401
    try:
        slot = int(request.args.get('slot', '-1'))
    except (TypeError, ValueError):
        slot = -1
    id_esclavo = (request.args.get('id_esclavo') or "").strip()
    cargas = obtener_auditoria_cargas_slot(slot, id_esclavo)
    return jsonify({"dashboard": calcular_dashboard_cargas(cargas), "cargas": cargas})

@app.route('/historial_excel')
def excel_dl():
    if check_auth(request)[0] == 'none': return "401", 401
    inicio = (request.args.get('inicio') or "").strip()
    fin = (request.args.get('fin') or "").strip()
    usuario = (request.args.get('usuario') or "").strip()
    maquina = (request.args.get('maquina') or "").strip()
    rango = (request.args.get('rango') or "").strip()
    
    where = []
    params =[]
    
    if usuario:
        where.append("usuario=?")
        params.append(usuario)
    if maquina:
        where.append("id_maquina=?")
        params.append(maquina)
        
    sql = 'SELECT fecha_hora, usuario, id_maquina, ip, monto FROM transacciones'
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    
    with db_lock:
        conn=sqlite3.connect(DB_NAME, timeout=10); c=conn.cursor()
        c.execute(sql, tuple(params))
        rows = c.fetchall(); conn.close()
        
    d_start = None
    if inicio:
        try:
            d_start = datetime.strptime(inicio, "%Y-%m-%d")
        except ValueError:
            d_start = None
    d_end = None
    if fin:
        try:
            d_end = datetime.strptime(fin, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError:
            d_end = None

    filtered_rows = []
    for row in rows:
        fh = parse_fecha_log(row[0])
        if not fh: continue
        
        ok = True
        if d_start and fh < d_start: ok = False
        if d_end and fh > d_end: ok = False
        if rango and not fecha_en_rango(row[0], rango): ok = False
            
        if ok:
            filtered_rows.append(row)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    ws.append(["Fecha y Hora", "Usuario", "Maquina", "IP", "Monto"])
    for row in filtered_rows: ws.append(row)
        
    filename = "Reporte_Auditoria.xlsx"
    wb.save(filename)
    return send_file(filename, as_attachment=True)

@app.route('/admin_log_excel')
def admin_log_excel():
    if check_auth(request)[0] != 'admin': return "403", 403
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, usuario, accion, detalle, ip FROM acciones_admin ORDER BY id DESC")
        rows = c.fetchall()
        conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log Administrativo"
    ws.append(["Fecha y Hora", "Usuario", "Accion", "Detalle", "IP"])
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)
    filename = os.path.join(BASE_DIR, "Log_Administrativo.xlsx")
    wb.save(filename)
    registrar_accion_admin(session.get('usuario'), "DESCARGAR_LOG_ADMIN", "Excel log administrativo", request.remote_addr)
    return send_file(filename, as_attachment=True)

@app.route('/system_log_excel')
def system_log_excel():
    if check_auth(request)[0] != 'admin': return "403", 403
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT fecha_hora, tipo, detalle FROM system_logs ORDER BY id DESC")
        rows = c.fetchall()
        conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log Sistema"
    ws.append(["Fecha y Hora", "Tipo", "Detalle"])
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 80)
    filename = os.path.join(BASE_DIR, "Log_Sistema.xlsx")
    wb.save(filename)
    registrar_accion_admin(session.get('usuario'), "DESCARGAR_LOG_SISTEMA", "Excel log sistema", request.remote_addr)
    return send_file(filename, as_attachment=True)

@app.route('/api/backup_forzado', methods=['POST'])
def backup_forzado():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    crear_backup_db()
    registrar_accion_admin(usr, "BACKUP_FORZADO", "Backup manual solicitado", request.remote_addr)
    return "Backup creado"

@app.route('/slot_log_excel')
def slot_log_excel():
    if check_auth(request)[0] == 'none': return "401", 401
    try:
        slot = int(request.args.get('slot', '-1'))
    except (TypeError, ValueError):
        return "Slot invalido", 400
    id_esclavo = (request.args.get('id_esclavo') or "").strip()
    if not id_esclavo and (slot < 0 or slot >= MAX_CLIENTS):
        return "Slot invalido", 400

    where =[]
    params =[]
    if id_esclavo:
        where.append("id_esclavo=?")
        params.append(id_esclavo)
    else:
        where.append("slot=?")
        params.append(slot)

    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute('''SELECT fecha_hora, slot, id_esclavo, nombre_maquina, ip, tipo, usuario, detalle, monto, contadores
                     FROM logs_slot WHERE ''' + " AND ".join(where) + ''' ORDER BY id DESC''', tuple(params))
        rows = c.fetchall()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Log Slot {slot}"
    ws.append(["Fecha y Hora", "Slot", "ID Esclavo", "Maquina", "IP", "Tipo", "Usuario", "Detalle", "Monto", "Contadores"])
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)

    etiqueta = nombre_archivo_seguro(id_esclavo if id_esclavo else str(slot))
    filename = os.path.join(BASE_DIR, f"Log_Slot_{etiqueta}.xlsx")
    wb.save(filename)
    return send_file(filename, as_attachment=True)

@app.route('/contaduria_slot_excel')
def contaduria_slot_excel():
    if check_auth(request)[0] == 'none': return "401", 401
    try:
        slot = int(request.args.get('slot', '-1'))
    except (TypeError, ValueError):
        return "Slot invalido", 400
    id_esclavo = (request.args.get('id_esclavo') or "").strip()
    if not id_esclavo and (slot < 0 or slot >= MAX_CLIENTS):
        return "Slot invalido", 400

    cargas = obtener_auditoria_cargas_slot(slot, id_esclavo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Contaduria Slot {slot}"
    ws.append(["Fecha y Hora", "Usuario", "Maquina", "IP", "Monto", "Estado"])
    for carga in cargas:
        ws.append([carga["fecha"], carga["usr"], carga["maq"], carga["ip"], carga["monto"], carga["estado"]])
    if not cargas:
        ws.append(["Sin cargas enviadas para este slot/id esclavo", "", "", "", "", ""])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    etiqueta = nombre_archivo_seguro(id_esclavo if id_esclavo else str(slot))
    filename = os.path.join(BASE_DIR, f"Contaduria_Cargas_Slot_{etiqueta}.xlsx")
    wb.save(filename)
    return send_file(filename, as_attachment=True)

@app.route('/api/limpiar_historial', methods=['POST'])
def clr_aud():
    global ultimo_cambio_db
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    if rol != 'admin': return "403", 403
    with db_lock:
        conn=sqlite3.connect(DB_NAME, timeout=10); c=conn.cursor(); c.execute('DELETE FROM transacciones'); conn.commit(); conn.close()
    ultimo_cambio_db = time.time()
    registrar_accion_admin(usr, "VACIAR_HISTORIAL", "DELETE transacciones", request.remote_addr)
    return "OK"

@app.route('/api/solicitar_limite', methods=['POST'])
def solicitar_limite():
    rol, usr, _ = check_auth(request)
    if rol == "none": return "401", 401
    monto = float(request.form.get('monto', 0))
    fh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("INSERT INTO solicitudes_limite (fecha_hora, usuario, monto, estado) VALUES (?, ?, ?, 'PENDIENTE')", (fh, usr, monto))
        conn.commit()
        conn.close()
    return "Solicitud enviada al administrador"

@app.route('/api/solicitudes')
def get_solicitudes():
    if check_auth(request)[0] != 'admin': return "403", 403
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT id, fecha_hora, usuario, monto FROM solicitudes_limite WHERE estado='PENDIENTE'")
        rows = c.fetchall()
        conn.close()
    return jsonify([{"id": r[0], "fecha": r[1], "usuario": r[2], "monto": r[3]} for r in rows])

@app.route('/api/solicitudes/aprobar', methods=['POST'])
def aprobar_solicitud():
    if check_auth(request)[0] != 'admin': return "403", 403
    sid = request.form.get('id')
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT usuario, monto FROM solicitudes_limite WHERE id=?", (sid,))
        req = c.fetchone()
        if req:
            usr, monto_solicitado = req[0], float(req[1])
            
            # Buscamos el rol del usuario
            c.execute("SELECT rol FROM usuarios WHERE username=?", (usr,))
            rol_row = c.fetchone()
            rol = rol_row[0] if rol_row else "operador"
            
            # Obtenemos el limite actual (usuario o rol)
            # Nota: No usamos obtener_limites_usuario directamente aqui por el db_lock (evitar re-entry si no fuera RLock, aunque lo es)
            c.execute("SELECT max_operacion, max_diario FROM limites_usuario WHERE username=?", (usr,))
            row = c.fetchone()
            cur_op, cur_dia = 0.0, 0.0
            if row:
                cur_op = float(row[0]) if row[0] is not None else 0.0
                cur_dia = float(row[1]) if row[1] is not None else 0.0
            else:
                c.execute("SELECT max_operacion, max_diario FROM limites_rol WHERE rol=?", (rol,))
                row_rol = c.fetchone()
                if row_rol:
                    cur_op = float(row_rol[0]) if row_rol[0] is not None else 0.0
                    cur_dia = float(row_rol[1]) if row_rol[1] is not None else 0.0
            
            nuevo_limite_dia = cur_dia + monto_solicitado
            
            c.execute("""INSERT INTO limites_usuario (username, max_operacion, max_diario) VALUES (?, ?, ?)
                         ON CONFLICT(username) DO UPDATE SET max_diario=excluded.max_diario""", (usr, cur_op, nuevo_limite_dia))
            
            c.execute("UPDATE solicitudes_limite SET estado='APROBADA' WHERE id=?", (sid,))
            conn.commit()
            registrar_accion_admin(session.get('usuario'), "APROBAR_SOLICITUD", f"Usuario {usr} +${monto_solicitado} (Nuevo total diario: {nuevo_limite_dia})", request.remote_addr)
        conn.close()
    return "Solicitud aprobada"

@app.route('/api/limites_rol')
def get_limites_rol():
    if check_auth(request)[0] != 'admin': return "403", 403
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("SELECT rol, max_operacion, max_diario FROM limites_rol")
        rows = c.fetchall()
        conn.close()
    return jsonify([{"rol": r[0], "max_operacion": r[1], "max_diario": r[2]} for r in rows])

@app.route('/api/limites_rol/edit', methods=['POST'])
def edit_limite_rol():
    rol_usuario, usr, _ = check_auth(request)
    if rol_usuario == "none": return "401", 401
    if rol_usuario != 'admin': return "403", 403
    rol = normalizar_texto_corto(request.form.get('rol'), 20)
    m_op = parsear_limite_decimal(request.form.get('max_operacion', 0))
    m_dia = parsear_limite_decimal(request.form.get('max_diario', 0))
    if rol not in ROLES_VALIDOS or m_op is None or m_dia is None:
        return "Parametros invalidos", 400
    with db_lock:
        conn = sqlite3.connect(DB_NAME, timeout=10)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO limites_rol VALUES (?, ?, ?)", (rol, m_op, m_dia))
        conn.commit()
        conn.close()
    registrar_accion_admin(usr, "EDITAR_LIMITE_ROL", f"Rol {rol}: OP {m_op}, DIA {m_dia}", request.remote_addr)
    return "Límite actualizado"

# ====================================================================
# RED TCP/UDP Y PARSEO DE SAS
# ====================================================================
def tcp_handler(sock, slot):
    sock.settimeout(1.0); buf = ""; u_pi = time.time(); ultimo_pong[slot] = time.time()
    try:
        sock.sendall(b"PING\r\n")
    except Exception as e:
        registrar_log_slot(slot, "ERROR", f"No se pudo enviar PING inicial: {e}")
    while True:
        if time.time() - u_pi > 3.0:
            try:
                sock.sendall(b"PING\r\n")
                u_pi = time.time()
            except Exception as e:
                registrar_log_slot(slot, "ERROR", f"No se pudo enviar PING: {e}")
                break
        try:
            raw_data = sock.recv(1024)
            if not raw_data: break
            try:
                ch = raw_data.decode('utf-8', errors='ignore')
            except Exception:
                continue
            buf += ch
            while '\n' in buf:
                lin, buf = buf.split('\n', 1); dat = lin.strip()
                if not dat: continue
                ultimo_pong[slot] = time.time()
                
                if dat.startswith("PONG"):
                    p = dat.split('|')
                    sas_anterior = sas_conectado[slot]
                    if len(p) > 1: sas_conectado[slot] = (p[1] == "1")
                    else: sas_conectado[slot] = False
                    if len(p) > 2 and p[2]: heartbeat_info[slot]["uptime"] = p[2]
                    if len(p) > 3 and p[3]: heartbeat_info[slot]["firmware"] = p[3]
                    if len(p) > 4 and p[4]: heartbeat_info[slot]["ultimo_error"] = p[4]
                    if sas_anterior and not sas_conectado[slot]:
                        ultima_sas_desconexion[slot] = time.time()
                    elif sas_conectado[slot]:
                        ultima_sas_desconexion[slot] = 0.0
                
                elif dat.startswith("REGISTRO|"): 
                    p = dat.split('|'); macs_esclavos[slot] = p[2] if len(p)>=3 else "Desc"; ultimos_eventos[slot] = "Conectado"
                    registrar_log_slot(slot, "EVENTO", f"ESP32 conectado. MAC: {macs_esclavos[slot]}")
                
                elif "|METERS|" in dat:
                    p = dat.split("|METERS|")
                    if len(p) > 1: 
                        ultimos_contadores[slot] = p[1].strip() 
                        ultimos_eventos[slot] = "✅ Contadores leídos"
                        registrar_log_slot(slot, "CONTADORES", "Contadores leidos correctamente", contadores=ultimos_contadores[slot])
                    
                elif "[+] CONFIRMADO" in dat or "[+] NOTIFICACIÓN" in dat:
                    ultimos_eventos[slot] = f"Exito: ${ultimo_monto[slot]}"
                    if "[+] CONFIRMADO" in dat:
                        registrar_historial(f"{nombres_esclavos[slot]} ({ids_clientes[slot]})", ips_esclavos[slot], ultimo_monto[slot], ultimo_usuario[slot])
                        registrar_log_slot(slot, "CARGA_CONFIRMADA", dat, ultimo_usuario[slot], ultimo_monto[slot], operation_id=ultimo_operation_id[slot])
                    else:
                        registrar_log_slot(slot, "EVENTO", dat, ultimo_usuario[slot], ultimo_monto[slot], operation_id=ultimo_operation_id[slot])
                else:
                    # Limpiamos prefijo ID:XXXX| para la UI
                    msg_ui = dat
                    if dat.startswith("ID:") and "|" in dat:
                        msg_ui = dat.split("|", 1)[1].strip()
                    
                    ultimos_eventos[slot] = msg_ui
                    tipo_log = "ERROR" if any(x in dat.upper() for x in ["ERROR", "BLOQUEO", "[X]", "FALLA", "TIMEOUT"]) else "EVENTO"
                    registrar_log_slot(slot, tipo_log, dat)
        except socket.timeout:
            pass
        except Exception as e:
            registrar_log_slot(slot, "ERROR", f"Conexion TCP interrumpida: {e}")
            break
        if time.time() - ultimo_pong[slot] > 15: 
            registrar_log_slot(slot, "ERROR", "Timeout de respuesta (Slot ocupado o desconectado)")
            break
    registrar_log_slot(slot, "EVENTO", "Slot desconectado")
    ultima_desconexion[slot] = time.time()
    clientes_tcp[slot] = None; sas_conectado[slot] = False; sock.close()

def network_loop():
    udp = socket.socket(socket.AF_INET, socket.SOCK_DGRAM); udp.bind(('0.0.0.0', PUERTO_UDP)); udp.setblocking(False)
    while True:
        time.sleep(0.1)
        try:
            d, a = udp.recvfrom(1024); m = d.decode().strip()
            if m.startswith("SAS_BEACON|"):
                hid = m[11:]; s = -1
                for j in range(MAX_CLIENTS):
                    if ids_clientes[j] == hid: s = j; break
                if s == -1:
                    for j in range(MAX_CLIENTS):
                        if ids_clientes[j] == "": s = j; break
                if s != -1 and not clientes_tcp[s]:
                    try:
                        t = socket.socket(socket.AF_INET, socket.SOCK_STREAM); t.settimeout(2.0); t.connect((a[0], 8080))
                        ids_clientes[s], ips_esclavos[s], ultimo_pong[s], clientes_tcp[s] = hid, a[0], time.time(), t
                        ultima_conexion[s] = time.time()
                        sas_conectado[s] = False 
                        ultima_sas_desconexion[s] = time.time()
                        with db_lock:
                            conn=sqlite3.connect(DB_NAME, timeout=10); c=conn.cursor(); c.execute("SELECT nombre FROM maquinas WHERE id_esclavo=?", (hid,)); r=c.fetchone(); conn.close()
                        nombres_esclavos[s] = r[0] if r else f"Maquina {s}"
                        threading.Thread(target=tcp_handler, args=(t, s), daemon=True).start()
                    except Exception as e:
                        print(f"[-] Error conectando slot {hid} en {a[0]}: {e}")
        except BlockingIOError:
            pass
        except UnicodeDecodeError as e:
            print(f"[-] Beacon UDP invalido: {e}")
        except Exception as e:
            print(f"[-] Error en network_loop: {e}")

if __name__ == '__main__':
    init_db()
    registrar_arranque_sistema()
    atexit.register(registrar_apagado_limpio)
    crear_backup_db()
    threading.Thread(target=backup_loop, daemon=True).start()
    threading.Thread(target=network_loop, daemon=True).start()

    # Hardware opcional: inicializar y arrancar threads solo si está disponible
    if _init_dht():
        threading.Thread(target=_read_dht_loop, daemon=True).start()
    if _init_oled():
        threading.Thread(target=_oled_display_loop, daemon=True).start()

    app.run(host='0.0.0.0', port=PUERTO_WEB)
